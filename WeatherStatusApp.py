# Import relevant modules
import time
import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl


keyWord = "weather"
def startBrowser():
    '''
    This method is used to launch the chrome browser, opens this url https://www.google.com/ and accepts google cookies.
    :return:
    '''
    PATH = "chromedriver.exe"
    global driver
    driver = webdriver.Chrome(PATH)
    pathToTheGoogle = 'https://www.google.com/'
    driver.get(pathToTheGoogle)
    cookiesBoxLocator = '//button[@id="L2AGLb"]'
    cookiesBox = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, cookiesBoxLocator)))
    cookiesBox.click()


def searchOnGoogle():
    '''
    This method is used to start the keyword search.
    :param keyWord: in our case, for this program, the key word is "weather".
    :return:
    '''
    searchBoxLocator = '//input[@class="gLFyf gsfi"]'
    searchBox = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, searchBoxLocator)))
    searchBox.click()
    searchBox.send_keys(keyWord)
    searchBox.submit()


def getWeather():
    '''
    This method is used to get all the information for the current day.
    information = time, location, day, weather status, temperature, rainfall, moisture and wind speed
    :return:
    '''
    temperatureBoxLocator = "//span[@id='wob_tm']"
    rainfallBoxLocator = "// span[@id='wob_pp']"
    moistureBoxLocator = "//span[@id='wob_hm']"
    windBoxLocator = "//span[@id='wob_ws']"
    locationBoxLocation = "//div[@id='wob_loc']"
    dayBoxLocator = "//div[@id ='wob_dts']"
    statusBoxLocator = "//div[@id='wob_dcp']"
    temperatureBox = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, temperatureBoxLocator)))
    temperatureBox.click()
    rainfallBox = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, rainfallBoxLocator)))
    rainfallBox.click()
    moistureBox = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, moistureBoxLocator)))
    moistureBox.click()
    windBox = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, windBoxLocator)))
    windBox.click()
    locationBox = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, locationBoxLocation)))
    locationBox.click()
    dayBox = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, dayBoxLocator)))
    dayBox.click()
    statusBox = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, statusBoxLocator)))
    statusBox.click()
    time = datetime.datetime.now().strftime("%Y:%m:%d - %H:%M:%S")
    file = openpyxl.load_workbook(path)
    sheet = file.active
    day, hour = dayBox.text.split(" ")
    sheet.cell(column=1, row=1, value= "Time")
    sheet.cell(column=2, row=1, value= "Location")
    sheet.cell(column=3, row=1, value= "Day")
    sheet.cell(column=4, row=1, value= "Status")
    sheet.cell(column=5, row=1, value= "Temperature")
    sheet.cell(column=6, row=1, value= "Rainfall")
    sheet.cell(column=7, row=1, value= "Moisture")
    sheet.cell(column=8, row=1, value= "Wind Speed")
    sheet.cell(column=1, row=sheet.max_row + 1, value= time)
    sheet.cell(column=2, row=sheet.max_row, value= locationBox.text)
    sheet.cell(column=3, row=sheet.max_row, value= day)
    sheet.cell(column=4, row=sheet.max_row, value= statusBox.text)
    sheet.cell(column=5, row=sheet.max_row, value= temperatureBox.text + "°C")
    sheet.cell(column=6, row=sheet.max_row, value= rainfallBox.text)
    sheet.cell(column=7, row=sheet.max_row, value= moistureBox.text)
    sheet.cell(column=8, row=sheet.max_row, value= windBox.text)
    file.save(path)
    print("Done...")
    driver.quit()


def mainloop(interval):
    '''
    This method combines all the other functions
    :param interval: the time between updates
    :return:
    '''
    while True:
        try:
            interval = int(interval)
            startBrowser()
            searchOnGoogle()
            getWeather()
            time.sleep(interval)
        except:
            return False


# you need to create an excel file to store the data
# pathModel: path = 'C:/Users/Liviu/Desktop/excelFile.xlsx'
path = input("The Path to the excel file: ")
interval = input("Select the time between updates: ")
mainloop(interval)
